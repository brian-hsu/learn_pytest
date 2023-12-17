import openpyxl
import datetime
from pprint import pprint

class DataDrivenLibrary:
    def __init__(self, file_name, sheet_name):
        self.workbook = openpyxl.load_workbook(file_name)
        self.sheet = self.workbook[sheet_name]

    def fetch_row_count(self):
        return self.sheet.max_row

    def fetch_column_count(self):
        return self.sheet.max_column

    def fetch_key_names(self):
        keys = []
        for col in range(1, self.fetch_column_count() + 1):
            keys.append(self.sheet.cell(row=1, column=col).value)
        return keys

    def update_request_with_data(self, row_number, json_request, key_list):
        for i in range(1, self.fetch_column_count() + 1):
            cell_value = self.sheet.cell(row=row_number, column=i).value

            # 檢查單元格值是否為datetime對象，如果是，則轉換為字串
            if isinstance(cell_value, datetime.datetime):
                cell_value = cell_value.strftime('%Y-%m-%d')

            json_request[key_list[i - 1]] = cell_value

        return json_request

# 使用範例
library = DataDrivenLibrary('./ts_data.xlsx', 'sheet1')
row_count = library.fetch_row_count()
column_count = library.fetch_column_count()

print(f"Row count: {row_count}, Column count: {column_count}")

keys = library.fetch_key_names()
json_request = {}  # 假定的JSON請求
updated_json = library.update_request_with_data(2, json_request, keys)

print(f"Updated JSON: {updated_json}")
pprint(updated_json)