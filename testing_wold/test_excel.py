import openpyxl
import requests
import datetime
from pathlib import Path


def test_add_students_from_excel():
    # 載入Excel檔案
    load_data = Path.cwd() / "testing_wold" / "ts_data.xlsx"
    workbook = openpyxl.load_workbook(load_data)
    sheet = workbook.active

    # API URL
    api_url = "https://httpbin.org/post"

    for i in range(2, sheet.max_row + 1):  # 跳過標題行，從第二行開始
        # 從Excel中讀取數據
        first_name = sheet.cell(row=i, column=1).value
        middle_name = sheet.cell(row=i, column=2).value
        last_name = sheet.cell(row=i, column=3).value
        date_of_birth = sheet.cell(row=i, column=4).value

        # 確保日期為字串格式
        if isinstance(date_of_birth, datetime.datetime):
            date_of_birth = date_of_birth.strftime('%Y-%m-%d')

        # 創建要發送的數據
        student_data = {
            "first_name": first_name,
            "middle_name": middle_name,
            "last_name": last_name,
            "date_of_birth": date_of_birth
        }

        # 發送POST請求
        response = requests.post(api_url, json=student_data)
        assert response.status_code == 200, "Response status code is not 200"

        print(f"Added student {first_name} {last_name} with status code {response.status_code}")

# 執行測試函數
test_add_students_from_excel()

